import openpyxl
from celery import shared_task
from django.http import HttpResponse,FileResponse
from .models import User, Export


@shared_task
def export_data_to_excel(): #filepath
    user_list_queryset = User.objects.filter(date_joined__lte = '2023-05-15').values('first_name','last_name','email', 'mobile')
    print(user_list_queryset)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        "attachment; filename=" + "user_list.xlsx"
    )
    print(response)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "User List"
    print(worksheet.title)
    # Define the titles for columns
    columns = [
        "Firstname",
        "Lastname",
        "Email",
        "Phone",
    ]
    row_num = 1
   
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
     # Iterate through all movies
    for user in user_list_queryset:    
        row_num += 1

    # Define the data for each cell in the row
        row = [
                user['first_name'],
                user['last_name'],
                user['email'],
                user['mobile'],
            ]
        print(row)
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    # for row_num, row_data in enumerate(data, start=1):
    #     for col_num, cell_data in enumerate(row_data, start=1):
    #         sheet.cell(row=row_num, column=col_num, value=cell_data)
    # response = FileResponse(open(filepath, 'w'))
    # response["Content-Disposition"] = (
    #     "attachment; filename=" + "user_list.xlsx"
    # )
    print(response,'---------------------------')
    workbook.save(response)
    # Export.objects.create(file = response)

    print(workbook)
