from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect
from django import forms
from django.contrib import messages
from django.contrib.auth import  authenticate,login,logout
from django.urls import reverse
from .forms import SignupForm 
# from accounts.models import User
from django.contrib.auth import get_user_model
from django.db.models.query_utils import Q
from django.contrib.auth.forms import PasswordResetForm
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.core.mail import send_mail, BadHeaderError
from django.http import HttpResponse,JsonResponse
from django.template.loader import render_to_string
from django.utils.crypto import get_random_string
from .helpus import Msghandler
import json
from geopy.geocoders import Nominatim
from django.views import View

# Create your views here.
User = get_user_model()

class Signup(View):
    def get(self,request):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.method == 'GET':
            latitude = float(request.GET.get('latitude'))
            longitude = float(request.GET.get('longitude'))
            context = self.get_address(latitude,longitude)    
            return JsonResponse(context)
        form = SignupForm()
        return render(request, "accounts/signup.html", {"form": form})
    
    def post(self,request):
        form = SignupForm(self.request.POST)
        print(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registration successful." )
            return redirect("home")
        messages.error(request, "Unsuccessful registration. Invalid information.")

    def get_address(self,latitude,longitude):
        dumps = []
        geolocator = Nominatim(user_agent="accounts")
        a = list()
        a.append(latitude)
        a.append(longitude)
        location = geolocator.reverse(a)
        dumps.append(str(location))
        context = {
            'address':dumps
        }
        return context


class SignIn(View):
    def get(self,request):
        return render(request, "accounts/signin.html")
    
    def post(self,request):
        resp = {'status':'failed'}
        username = request.POST["email"]
        password = request.POST["password"]
        try:
            new_user = User.objects.get(mobile = username)
            user = authenticate(username=new_user, password=password)
        except:
            user = authenticate(username=username, password=password)
        if user:
            login(request,user)
                        
            resp['status'] = 'success'

        else:
            resp['status'] = 'failed'
        return HttpResponse(json.dumps(resp),content_type = 'application/json')


def sign_out(request):
    logout(request)
    messages.info(request, "Logged out successfully!")
    return redirect("home")


global_var = 0

def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        print(request.POST)
        global global_var
        
        otp = get_random_string(6, allowed_chars='123456789')
        global_var = int(otp)
        
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            user = User.objects.get(email = data)
            mobile = user.mobile

            # associated_users = User.objects.filter(Q(email=data))   
            # if associated_users.exists():       
                # for user in associated_users:
                    # subject = "Password Reset Requested"
                    # email_template_name = "accounts/password_reset_email.txt"
            mobile  = '+91'+ mobile
            obj = Msghandler(mobile,otp)
            obj.send_otp()
            # c = {
            #     "otp":otp
            # }
                    # email = render_to_string(email_template_name, c)
                    # response = fast2sms.requests.request('POST',url=fast2sms.url, data=fast2sms.payload , headers=fast2sms.headers)
                    # print(email,"____________________________________")
                    # print(response)
                    # try:
                    #     send_mail(subject, email, 'admin@example.com' , [user.email], fail_silently=False)
                    # except BadHeaderError:
                    #     return HttpResponse('Invalid header found.')
                    
                    # context ={
                    #     "mail":data
                    # }
                    
            return render (request,'accounts/password_reset_done.html')#, context)
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="accounts/password_reset.html", context={"password_reset_form":password_reset_form})

global_mail = ''

def resetpassword(request):
    if request.method == "POST":
        data = request.POST.get('mail')
        print(data)
        global global_mail
        global_mail = data
 
        myotp = request.POST.get('otp')
        print(myotp)
        print(global_var)

        if myotp == str(global_var):
            
            context ={
                        "mail":data
                    }
            return render(request,"accounts/password_reset_confirm.html",context)
        else:
            return redirect('password_reset_done')
        
    return render(request,"accounts/password_reset_confirm.html")

def resetpassword_complete(request):
    if request.method == "POST":
        mail = request.POST.get('mail')
        pass1 = request.POST.get('password')
        pass2 = request.POST.get('password1')
        user = User.objects.get(email = mail)
        print(user)
        if pass1 == pass2:
            user.set_password(pass1)
            user.save()
            
            return render(request,"accounts/password_reset_complete.html")
        return redirect('password_reset_confirm')
    

def user_list(request):
    user_list = User.objects.filter(date_joined__lte = '2023-5-15')
    print(user_list)
    return HttpResponse('userlist exported ')

from .tasks import export_data_to_excel
from django.http import HttpResponse
from openpyxl import load_workbook

def export_view(request):
    # filepath = f'/home/cis/Truevalue/truevalue/user_list.xlsx'
    export_data_to_excel.delay() #filepath


    # Load the workbook
    workbook = load_workbook('/xl/workbook.xml/user_list_queryset.xlsx')
    print('-----------------------')
    # Get the list of sheet names
    sheet_names = workbook.sheetnames

    # Print the sheet names
    for sheet_name in sheet_names:
        print(sheet_name)
    return HttpResponse('Export task has been triggered.')